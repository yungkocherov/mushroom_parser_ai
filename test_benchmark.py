"""
Бенчмарк моделей на тестовых фото.
Каждый запуск добавляет столбец с результатами текущей модели.

Использование:
  1. Поменяй MODEL на нужную
  2. Запусти: python test_benchmark.py
  3. Повтори для другой модели
  4. Открой data/test_results.json

Запуск: python test_benchmark.py
"""

import requests
import base64
import sys
import time
import json
import os

sys.stdout.reconfigure(encoding='utf-8')

# === НАСТРОЙКИ — МЕНЯЙ ПЕРЕД ЗАПУСКОМ ===
API_URL = "http://localhost:1234/v1/chat/completions"
MODEL = "google/gemma-4-e4b"

PHOTOS_FILE  = "data/test_photos.json"
RESULTS_FILE = "data/test_results.json"

PROMPT = """/no_think
Classify mushrooms in this photo into one of these groups and count them.
JSON only: [{"species":"group","count":N}]
No mushrooms: [{"species":"none","count":0}]
Basket=30-50, handful=5-10.

Groups:
- chanterelle (Cantharellus, any chanterelle species)
- bolete (Boletus, Leccinum, Suillus, Xerocomus - any tube mushroom with sponge under cap)
- morel (Morchella, Gyromitra, Verpa - wrinkled/brain-like cap, spring mushrooms)
- honey_fungus (Armillaria, Kuehneromyces - clusters on wood)
- other (any mushroom not matching above groups)"""


def ask_model(img_b64):
    import re
    try:
        resp = requests.post(API_URL, json={
            "model": MODEL,
            "messages": [{"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                {"type": "text", "text": PROMPT},
            ]}],
            "temperature": 0.1,
            "max_tokens": 1000,  # больше токенов для thinking
        }, timeout=120)
        if resp.status_code != 200:
            return f"ERROR:{resp.status_code}"
        msg = resp.json()["choices"][0]["message"]
        content = msg.get("content", "")
        if not content and msg.get("reasoning_content"):
            content = msg["reasoning_content"]
        # Ищем JSON массив в конце ответа (после thinking)
        matches = re.findall(r"\[.*?\]", content, re.DOTALL)
        if matches:
            return matches[-1].strip().replace("\n", " ")[:300]
        return content.strip().replace("\n", " ")[:300]
    except Exception as e:
        return f"ERROR:{str(e)[:80]}"


def main():
    # Загружаем URL фото
    with open(PHOTOS_FILE, encoding="utf-8") as f:
        urls = json.load(f)

    # Загружаем предыдущие результаты
    if os.path.exists(RESULTS_FILE):
        with open(RESULTS_FILE, encoding="utf-8") as f:
            results = json.load(f)
    else:
        results = {"urls": urls, "models": {}}

    col_name = MODEL.replace("/", "_")
    print(f"Модель: {MODEL}")
    print(f"Столбец: {col_name}")
    print(f"Фото: {len(urls)}")
    print()

    model_results = []
    model_times = []
    total_time = 0

    for i, url in enumerate(urls):
        try:
            img = requests.get(url, timeout=15).content
        except Exception:
            model_results.append("SKIP")
            model_times.append(0)
            print(f"{i+1:2d}. SKIP (download error)")
            continue
        if len(img) < 1000:
            model_results.append("SKIP")
            model_times.append(0)
            print(f"{i+1:2d}. SKIP (bad image, {len(img)} bytes)")
            continue

        img_b64 = base64.b64encode(img).decode()

        start = time.time()
        answer = ask_model(img_b64)
        elapsed = time.time() - start
        total_time += elapsed

        model_results.append(answer)
        model_times.append(round(elapsed, 1))
        print(f"{i+1:2d}. ({elapsed:.1f}s) {answer[:120]}")

    results["models"][col_name] = model_results
    results["times"] = results.get("times", {})
    results["times"][col_name] = model_times

    # Сохраняем
    with open(RESULTS_FILE, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    avg = total_time / len(urls) if urls else 0
    print(f"\nСреднее: {avg:.1f}s/фото")
    print(f"Результат сохранён в {RESULTS_FILE}")
    print(f"Моделей протестировано: {len(results['models'])}")

    # Сохраняем Excel для удобного просмотра
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    xlsx_file = RESULTS_FILE.replace(".json", ".xlsx")
    model_names = list(results["models"].keys())
    times_data = results.get("times", {})

    def parse_answer(raw):
        raw = re.sub(r'```json\s*', '', raw).replace('```', '').strip()
        try:
            items = json.loads(raw) if raw.startswith('[') else [json.loads(raw)]
            parts = []
            for item in items:
                sp = item.get("species", "?")
                cnt = item.get("count", "?")
                if sp not in ("none", "нет_грибов"):
                    parts.append(f"{sp}={cnt}")
            return ", ".join(parts) if parts else "нет грибов"
        except:
            if "none" in raw.lower():
                return "нет грибов"
            return raw[:80]

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark"

    ground_truth = results.get("ground_truth", [])

    # Заголовок
    headers = ["#", "URL", "РЕАЛЬНОСТЬ"]
    for m in model_names:
        headers.extend([m, f"время"])
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    # Данные
    for i, url in enumerate(results["urls"]):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        # URL как гиперссылка
        cell = ws.cell(row=row, column=2, value="фото")
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")
        # Ground truth
        gt = ground_truth[i] if i < len(ground_truth) else ""
        gt_cell = ws.cell(row=row, column=3, value=gt)
        gt_cell.font = Font(bold=True, color="006600")

        col_idx = 4
        for m in model_names:
            vals = results["models"][m]
            raw = vals[i] if i < len(vals) else ""
            ws.cell(row=row, column=col_idx, value=parse_answer(raw))
            t = times_data.get(m, [])
            ws.cell(row=row, column=col_idx + 1, value=t[i] if i < len(t) else 0)
            col_idx += 2

    # Ширина столбцов
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 35
    for c in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 35

    wb.save(xlsx_file)
    print(f"Excel: {xlsx_file}")


if __name__ == "__main__":
    main()
